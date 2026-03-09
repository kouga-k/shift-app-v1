import streamlit as st
import pandas as pd
from ortools.sat.python import cp_model
import io
import jpholiday
import datetime
import random
import calendar
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── ログイン管理 ──
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

if not st.session_state["logged_in"]:
    st.set_page_config(page_title="ログイン", layout="centered")
    st.title("🔐 ログイン")
    user_id  = st.text_input("ID")
    password = st.text_input("PASS", type="password")
    if st.button("ログイン"):
        if user_id == "admin" and password == "1234":
            st.session_state["logged_in"] = True
            st.rerun()
        else:
            st.error("IDまたはパスワードが違います")
    st.stop()

st.set_page_config(page_title="自動シフト作成アプリ", layout="wide")
st.title("📅 シフト自動作成")
st.write("希望休・夜勤ルール・役割条件を考慮して、最適なシフトを自動で作成します。")

if 'needs_compromise' not in st.session_state:
    st.session_state.needs_compromise = False

st.write("---")
today = datetime.date.today()
col_y, col_m = st.columns(2)
with col_y: target_year = st.selectbox("作成年", [today.year, today.year + 1], index=0)
with col_m: target_month = st.selectbox("作成月", list(range(1, 13)), index=(today.month % 12))
st.write("---")

uploaded_file = st.file_uploader("エクセルファイル (.xlsx) を選択", type=["xlsx"])

if uploaded_file:
    try:
        df_staff = pd.read_excel(uploaded_file, sheet_name="スタッフ設定")
        df_history = pd.read_excel(uploaded_file, sheet_name="希望休・前月履歴")
        df_req = pd.read_excel(uploaded_file, sheet_name="日別設定")

        staff_names = df_staff["スタッフ名"].dropna().tolist()
        num_staff = len(staff_names)

        def get_staff_col(col_name, default_val, is_int=False):
            res = []
            for i in range(num_staff):
                if col_name in df_staff.columns and pd.notna(df_staff[col_name].iloc[i]):
                    val = df_staff[col_name].iloc[i]
                    res.append(int(val) if is_int else str(val).strip())
                else:
                    res.append(default_val)
            return res

        staff_roles = get_staff_col("役割", "一般")

        # ── 公休数：週休2日制 → 2月=8日、それ以外=9日を自動設定（手入力が優先）──
        _auto_off = 8 if target_month == 2 else 9
        staff_off_days = get_staff_col("公休数", _auto_off, is_int=True)
        staff_night_ok = get_staff_col("夜勤可否", "〇")
        staff_overtime_ok = get_staff_col("残業可否", "〇")
        staff_part_shifts = get_staff_col("パート", "")
        staff_night_limits = [0 if ok == "×" else int(v) if pd.notna(v) else 10 for ok, v in zip(staff_night_ok, get_staff_col("夜勤上限", 10, is_int=True))]
        staff_min_normal_a = get_staff_col("定時確保数", 2, is_int=True)
        staff_sun_d = ["×" if ok == "×" else v for ok, v in zip(staff_night_ok, get_staff_col("日曜Dカウント", "〇"))]
        staff_sun_e = ["×" if ok == "×" else v for ok, v in zip(staff_night_ok, get_staff_col("日曜Eカウント", "〇"))]

        # 有休・夏冬休関連
        staff_join_month   = get_staff_col("入職月", 0, is_int=True)          # 入職月（1〜12）
        staff_paid_given   = get_staff_col("有休付与日数", 10, is_int=True)    # 今年度付与日数
        staff_paid_taken   = get_staff_col("有休取得済", 0, is_int=True)       # 今年度取得済累計
        staff_summer_given = get_staff_col("夏季休暇付与", 3, is_int=True)     # 夏休付与日数（7〜9月で3日）
        staff_summer_taken = get_staff_col("夏季休暇取得済", 0, is_int=True)   # 夏休取得済
        staff_winter_given = get_staff_col("冬季休暇付与", 4, is_int=True)     # 冬休付与日数（12〜2月で4日）
        staff_winter_taken = get_staff_col("冬季休暇取得済", 0, is_int=True)   # 冬休取得済

        # 前月の公休実績（年間管理シートから）→ 繰り越しチェック用
        staff_prev_off_actual = []
        for e in range(num_staff):
            val = get_annual_val(staff_names[e], "前月公休実績", _auto_off)
            staff_prev_off_actual.append(int(val))

        # 休暇優先順位ルール
        # 公休(9日) ＞ 夏冬休 ＞ 年休
        # 夏休：7〜9月に3日、冬休：12〜2月に4日
        SUMMER_MONTHS = [7, 8, 9]
        WINTER_MONTHS = [12, 1, 2]
        SUMMER_TARGET = 3
        WINTER_TARGET = 4

        # 入職月別 義務取得日数テーブル（画像より）
        PAID_OBLIGATION_TABLE = {
            4: 2.5, 5: 2.5, 6: 2.0, 7: 1.5, 8: 1.0, 9: 0.5,
            10: 5.0, 11: 5.0, 12: 4.5, 1: 4.0, 2: 3.5, 3: 3.0
        }
        # 入職月 → 最初の有休付与月（入職6ヶ月後）
        def get_grant_month(join_month):
            if join_month == 0: return 0
            return ((join_month - 1 + 6) % 12) + 1

        staff_comp_lvl = []
        for i in range(num_staff):
            val = ""
            if "妥協優先度" in df_staff.columns and pd.notna(df_staff["妥協優先度"].iloc[i]):
                val = str(df_staff["妥協優先度"].iloc[i]).strip()
            elif "連勤妥協OK" in df_staff.columns and pd.notna(df_staff["連勤妥協OK"].iloc[i]):
                val = str(df_staff["連勤妥協OK"].iloc[i]).strip()
            if val in ["〇", "1", "1.0"]: staff_comp_lvl.append(1)
            elif val in ["2", "2.0"]: staff_comp_lvl.append(2)
            elif val in ["3", "3.0"]: staff_comp_lvl.append(3)
            else: staff_comp_lvl.append(0)

        date_columns = [col for col in df_req.columns if col != df_req.columns[0] and not str(col).startswith("Unnamed")]
        num_days = len(date_columns)

        def get_req_col(label, default_val, is_int=True):
            row = df_req[df_req.iloc[:, 0] == label]
            res = []
            for d in range(num_days):
                col_name = date_columns[d]
                try:
                    if not row.empty and col_name in df_req.columns:
                        val = row.iloc[0][col_name]
                        if pd.notna(val):
                            res.append(int(val) if is_int else str(val).strip())
                            continue
                except Exception:
                    pass
                res.append(default_val)
            return res

        day_req_list = get_req_col("日勤人数", 3)
        night_req_list = get_req_col("夜勤人数", 2)
        overtime_req_list = get_req_col("残業人数", 0)
        absolute_req_list = get_req_col("絶対確保", "", is_int=False)

        _weekday_map = ["月", "火", "水", "木", "金", "土", "日"]
        weekdays = []
        for _d_val in date_columns:
            try:
                _dt = datetime.date(target_year, target_month, int(_d_val))
                weekdays.append(_weekday_map[_dt.weekday()])
            except (ValueError, TypeError):
                weekdays.append("")

        # =============================================
        # 希望休情報を事前集計（カード生成・診断で共用）
        # =============================================
        def build_fixed_off_info():
            fixed_off = [0] * num_staff
            fixed_off_days_list = [[] for _ in range(num_staff)]
            fixed_off_per_day = [0] * num_days
            for e, staff_name in enumerate(staff_names):
                tr = df_history[df_history.iloc[:, 0] == staff_name]
                if not tr.empty:
                    history_today_cols = list(tr.columns[6:])
                    for d in range(num_days):
                        col_name = date_columns[d]
                        if col_name in history_today_cols and str(tr.iloc[0][col_name]).strip() == "公":
                            fixed_off[e] += 1
                            fixed_off_days_list[e].append(d)
                            fixed_off_per_day[d] += 1
            return fixed_off, fixed_off_days_list, fixed_off_per_day

        fixed_off, fixed_off_days_list, fixed_off_per_day = build_fixed_off_info()

        # =============================================
        # 📊 年間管理シート読み込み
        # =============================================
        try:
            df_annual = pd.read_excel(uploaded_file, sheet_name="年間管理")
        except Exception:
            # 年間管理シートがない場合は空のDataFrameを作成
            df_annual = pd.DataFrame(columns=[
                "スタッフ名", "入職月", "有休付与日数", "有休取得済(累計)", "有休残日数",
                "義務取得日数", "義務残日数", "夏季付与", "夏季取得済", "冬季付与", "冬季取得済",
                "夜勤累計", "残業累計", "連勤最大(過去最高)", "最終更新"
            ])

        def get_annual_val(name, col, default=0):
            """年間管理シートから特定スタッフの値を取得"""
            if df_annual.empty or col not in df_annual.columns:
                return default
            row = df_annual[df_annual["スタッフ名"] == name]
            if row.empty or pd.isna(row.iloc[0][col]):
                return default
            try:
                return float(row.iloc[0][col])
            except Exception:
                return default

        # =============================================
        # 📊 振り返りレポート生成関数
        # =============================================
        # 有休年度ルール：4月〜2月の間に年5日取得。3月は未取得分を強制消化。
        # 現在が何月かで「残り取得可能月数」と「必要ペース」を計算する。
        PAID_YEAR_TARGET = 5          # 年間義務取得日数
        # 4月=1ヶ月目、5月=2ヶ月目…2月=11ヶ月目、3月は強制消化月
        def months_in_period(m):
            """4月を起点にした経過月数（4月=1, 5月=2, …, 2月=11, 3月=12）"""
            return ((m - 4) % 12) + 1

        def remaining_months_to_feb(m):
            """現在月から2月末までの残り月数（3月シフト作成時は0）"""
            if m == 3: return 0
            period = months_in_period(m)
            return 11 - period  # 11ヶ月目(2月)まであと何ヶ月

        def build_review_report(df_fin, cols):
            """
            完成シフトdf_finから今月実績を集計し、以下を返す：
            ・有休：4〜2月で年5日取得の月別ペース促し
            ・夏休：7〜9月で3日取得促し
            ・冬休：12〜2月で4日取得促し
            ・公休：8日だった翌月は10日繰り越し警告
            ・優先順位：公休 ＞ 夏冬休 ＞ 年休
            """
            report_rows = []
            alerts       = []   # 黄色警告
            red_alerts   = []   # 赤色強制対応
            march_forced = []   # 3月強制消化

            remain_months = remaining_months_to_feb(target_month)
            is_march = (target_month == 3)

            # 今月の公休実績（シフト表から集計）→ 来月繰り越し判定用
            off_actual_this = {}

            for e in range(num_staff):
                name = staff_names[e]

                # ── 今月実績 ──
                night_this  = int(df_fin.loc[e, "夜勤(D)回数"]) if "夜勤(D)回数" in df_fin.columns else 0
                ot_this     = int(df_fin.loc[e, "残業(A残)回数"]) if "残業(A残)回数" in df_fin.columns else 0
                off_this    = int(df_fin.loc[e, "公休回数"]) if "公休回数" in df_fin.columns else _auto_off
                off_actual_this[name] = off_this

                # 連勤最大日数
                max_consec = 0; cur = 0
                for d in range(num_days):
                    v = str(df_fin.loc[e, cols[d]]) if d < len(cols) else ""
                    if v in ['A', 'A残'] or 'P' in v:
                        cur += 1; max_consec = max(max_consec, cur)
                    else:
                        cur = 0

                # ── 年間累計 ──
                night_cum       = int(get_annual_val(name, "夜勤累計", 0)) + night_this
                ot_cum          = int(get_annual_val(name, "残業累計", 0)) + ot_this
                max_consec_past = int(get_annual_val(name, "連勤最大(過去最高)", 0))
                max_consec_all  = max(max_consec_past, max_consec)

                # ────────────────────────────────────────
                # ① 公休チェック（最優先）
                # ────────────────────────────────────────
                prev_off = staff_prev_off_actual[e]   # 前月公休実績
                carryover_needed = 0
                off_judgement = []

                if prev_off < 9:
                    shortage = 9 - prev_off
                    carryover_needed = 9 + shortage  # 来月は9+不足分
                    off_judgement.append(f"⚠️ 前月公休{prev_off}日→今月{carryover_needed}日繰越必要")
                    alerts.append(
                        f"**{name}**：前月の公休が **{prev_off}日** でした（基準9日）。"
                        f"今月は **{carryover_needed}日** の公休が必要です。"
                        f"※公休不足は緊急時のみ。しっかり休養するよう声がけを。"
                    )
                if off_this < 9 and off_this > 0:
                    off_judgement.append(f"🟡 今月公休{off_this}日（基準9日）")
                    alerts.append(
                        f"**{name}**：今月の公休が **{off_this}日** です。"
                        f"週休2日制のため基準は9日です。緊急時以外はしっかり休養させてください。"
                        f"（来月は繰り越し{9+(9-off_this)}日が必要になります）"
                    )

                # ────────────────────────────────────────
                # ② 夏季休暇チェック（7〜9月）
                # ────────────────────────────────────────
                sum_given  = int(get_annual_val(name, "夏季付与", SUMMER_TARGET))
                sum_taken  = int(get_annual_val(name, "夏季取得済", staff_summer_taken[e]))
                sum_remain = max(0, sum_given - sum_taken)
                sum_judgement = []

                if target_month in SUMMER_MONTHS:
                    # 夏休期間中
                    months_left_summer = SUMMER_MONTHS[-1] - target_month  # 9月まであと何ヶ月
                    if sum_remain > 0:
                        if months_left_summer == 0:
                            # 9月末が最後のチャンス
                            sum_judgement.append(f"🔴 夏休{sum_remain}日未取得（今月が最終）")
                            red_alerts.append(
                                f"**{name}**：夏季休暇が **{sum_remain}日** 未取得です。"
                                f"9月が最終月のため、今月のシフトに必ず組み込んでください。"
                            )
                        else:
                            sum_judgement.append(f"🟡 夏休残{sum_remain}日（残{months_left_summer}ヶ月）")
                            alerts.append(
                                f"**{name}**：夏季休暇が残 **{sum_remain}日** です。"
                                f"9月末までに取得してください（残{months_left_summer}ヶ月）。"
                            )
                    else:
                        sum_judgement.append("🟢 夏休取得完了")
                elif target_month > 9:
                    if sum_remain > 0:
                        sum_judgement.append(f"⚠️ 夏休{sum_remain}日未消化（期間外）")

                # ────────────────────────────────────────
                # ③ 冬季休暇チェック（12〜2月）
                # ────────────────────────────────────────
                win_given  = int(get_annual_val(name, "冬季付与", WINTER_TARGET))
                win_taken  = int(get_annual_val(name, "冬季取得済", staff_winter_taken[e]))
                win_remain = max(0, win_given - win_taken)
                win_judgement = []

                if target_month in WINTER_MONTHS:
                    # 冬休期間中：12月=残2ヶ月、1月=残1ヶ月、2月=最終
                    months_left_winter = (2 - target_month) % 12  # 2月まであと何ヶ月（12月→2, 1月→1, 2月→0）
                    if target_month == 12: months_left_winter = 2
                    elif target_month == 1: months_left_winter = 1
                    else: months_left_winter = 0

                    if win_remain > 0:
                        if months_left_winter == 0:
                            win_judgement.append(f"🔴 冬休{win_remain}日未取得（今月が最終）")
                            red_alerts.append(
                                f"**{name}**：冬季休暇が **{win_remain}日** 未取得です。"
                                f"2月が最終月のため、今月のシフトに必ず組み込んでください。"
                            )
                        else:
                            win_judgement.append(f"🟡 冬休残{win_remain}日（残{months_left_winter}ヶ月）")
                            alerts.append(
                                f"**{name}**：冬季休暇が残 **{win_remain}日** です。"
                                f"2月末までに取得してください（残{months_left_winter}ヶ月）。"
                            )
                    else:
                        win_judgement.append("🟢 冬休取得完了")
                elif target_month > 2 and target_month < 12:
                    if win_remain > 0:
                        win_judgement.append(f"⚠️ 冬休{win_remain}日未消化")

                # ────────────────────────────────────────
                # ④ 有休チェック（4〜2月 年5日）※最低優先
                # ────────────────────────────────────────
                paid_given       = int(get_annual_val(name, "有休付与日数", staff_paid_given[e]))
                paid_taken_total = float(get_annual_val(name, "有休取得済(累計)", staff_paid_taken[e]))
                paid_remain      = max(0.0, paid_given - paid_taken_total)
                year_short       = max(0.0, PAID_YEAR_TARGET - paid_taken_total)

                join_m = staff_join_month[e]
                first_year_obligation = PAID_OBLIGATION_TABLE.get(join_m, 5.0) if join_m != 0 else 5.0
                first_year_short = max(0.0, first_year_obligation - paid_taken_total)

                paid_judgement = []

                if is_march and year_short > 0:
                    paid_judgement.append(f"🔴 3月強制消化{year_short}日")
                    red_alerts.append(
                        f"**{name}**：年間有休5日に対し **{year_short}日** 未達です。"
                        f"3月シフトに必ず有休を組み込んでください。"
                    )
                    march_forced.append({"name": name, "days": year_short})
                elif not is_march:
                    period_elapsed   = months_in_period(target_month)
                    expected_by_now  = round(PAID_YEAR_TARGET * period_elapsed / 11, 1)
                    if remain_months > 0:
                        needed_per_month = round(year_short / remain_months, 1) if year_short > 0 else 0
                        pace_msg = f"残{remain_months}ヶ月・残{year_short}日（月{needed_per_month}日ペース）"
                    else:
                        pace_msg = "2月末 達成済み" if year_short == 0 else f"2月末 {year_short}日未達"

                    if year_short == 0:
                        paid_judgement.append("🎉 有休5日達成")
                    elif paid_taken_total < expected_by_now - 1:
                        paid_judgement.append(f"🟡 有休ペース遅れ（{pace_msg}）")
                        alerts.append(
                            f"**{name}**：有休取得が遅れています。{pace_msg}。"
                            f"このまま進むと2月末に義務未達になる恐れがあります。"
                            f"公休・夏冬休の後に有休を計画的に配置してください。"
                        )
                    else:
                        paid_judgement.append(f"🟢 有休進行中（{pace_msg}）")

                if first_year_short > 0 and join_m != 0:
                    paid_judgement.append(f"⚠️ 初年度義務残{first_year_short}日")

                # ── 全判定まとめ ──
                judgements = off_judgement + sum_judgement + win_judgement + paid_judgement
                if max_consec >= 4:
                    judgements.append(f"⚠️ 最大{max_consec}連勤")
                if not judgements:
                    judgements.append("🟢 問題なし")

                report_rows.append({
                    "スタッフ名":     name,
                    "役割":           staff_roles[e],
                    "今月公休":       off_this,
                    "今月夜勤":       night_this,
                    "今月残業":       ot_this,
                    "今月最大連勤":   max_consec,
                    "夜勤累計(年)":   night_cum,
                    "残業累計(年)":   ot_cum,
                    "有休取得済":     paid_taken_total,
                    "有休残":         paid_remain,
                    "年間義務残":     year_short,
                    "夏休残":         sum_remain,
                    "冬休残":         win_remain,
                    "判定":           " / ".join(judgements),
                    # 年間管理更新用
                    "_night_cum":     night_cum,
                    "_ot_cum":        ot_cum,
                    "_max_consec":    max_consec_all,
                    "_paid_given":    paid_given,
                    "_paid_taken":    paid_taken_total,
                    "_paid_remain":   paid_remain,
                    "_year_short":    year_short,
                    "_sum_given":     sum_given,
                    "_sum_taken":     sum_taken,
                    "_win_given":     win_given,
                    "_win_taken":     win_taken,
                    "_off_this":      off_this,
                })

            df_report = pd.DataFrame(report_rows)
            return df_report, alerts, red_alerts, march_forced


        def build_annual_excel(df_report):
            """振り返りレポートをもとに年間管理シートを更新したExcelを返す"""
            update_month = f"{target_year}/{target_month:02d}"
            rows = []
            for _, r in df_report.iterrows():
                rows.append({
                    "スタッフ名":           r["スタッフ名"],
                    "入職月":               staff_join_month[staff_names.index(r["スタッフ名"])],
                    "有休付与日数":          r["_paid_given"],
                    "有休取得済(累計)":      r["_paid_taken"],
                    "有休残日数":            r["_paid_remain"],
                    "年間義務残日数":        r["_year_short"],
                    "夏季付与":             r["_sum_given"],
                    "夏季取得済":           r["_sum_taken"],
                    "冬季付与":             r["_win_given"],
                    "冬季取得済":           r["_win_taken"],
                    "夜勤累計":             r["_night_cum"],
                    "残業累計":             r["_ot_cum"],
                    "連勤最大(過去最高)":    r["_max_consec"],
                    "前月公休実績":          r["_off_this"],   # ← 来月の繰り越しチェック用
                    "最終更新":              update_month,
                })
            return pd.DataFrame(rows)




        def generate_compromise_cards(necessary):
            """
            必要な妥協案ごとに「誰が・何日・なぜ」を具体的に示すカードを生成する。
            """
            cards = []

            # ── カード0：平日・祝の日勤人数を-1 ──
            if necessary[0]:
                short_days = []
                for d in range(num_days):
                    if '日' not in weekdays[d] and '土' not in weekdays[d]:
                        blocked = len(fixed_off_days_list[e] for e in range(num_staff) if d in fixed_off_days_list[e])
                        # ブロック人数を正しく計算
                        blocked_count = sum(1 for e in range(num_staff) if d in fixed_off_days_list[e])
                        night_use = night_req_list[d]
                        e_use = night_req_list[d - 1] if d > 0 else 0
                        available = num_staff - blocked_count - night_use - e_use
                        if available < day_req_list[d]:
                            short_days.append(f"{date_columns[d]}日({weekdays[d]})")
                if short_days:
                    days_str = "・".join(short_days[:4]) + ("など" if len(short_days) > 4 else "")
                    detail = f"特に **{days_str}** で人手が不足する見込みです。設定人数より1名少ない状態を許容します。"
                else:
                    detail = "一部の平日・祝日で日勤人数が設定より1名少なくなることを許容します。"
                cards.append({
                    "id": 0,
                    "icon": "📉",
                    "title": "平日・祝の日勤を1名減で許容",
                    "detail": detail,
                    "impact": "軽微",
                    "impact_color": "#28a745",
                    "bg": "#f8fff8",
                    "border": "#28a745",
                })

            # ── カード1：4連勤 ──
            if necessary[1]:
                comp_staff = [staff_names[e] for e in range(num_staff) if staff_comp_lvl[e] > 0]
                if comp_staff:
                    names_str = "・".join(comp_staff)
                    detail = f"**{names_str}** さんに、月内で最大4日連続勤務をお願いする可能性があります（妥協優先度が設定されているスタッフのみ対象）。"
                else:
                    detail = "妥協優先度が設定されたスタッフに、4日連続勤務を許容します。"
                cards.append({
                    "id": 1,
                    "icon": "📅",
                    "title": "4連勤のお願い（対象者限定）",
                    "detail": detail,
                    "impact": "中程度",
                    "impact_color": "#fd7e14",
                    "bg": "#fffbf5",
                    "border": "#fd7e14",
                })

            # ── カード2：夜勤前3連日勤 ──
            if necessary[2]:
                comp_night_staff = [staff_names[e] for e in range(num_staff) if staff_comp_lvl[e] > 0 and staff_night_ok[e] != "×"]
                if comp_night_staff:
                    names_str = "・".join(comp_night_staff)
                    detail = f"**{names_str}** さんに、夜勤(D)の直前3日間を連続日勤にする場合があります。通常は夜勤前の連勤数を制限していますが、この制限を一部緩和します。"
                else:
                    detail = "妥協優先度が設定された夜勤可能スタッフに、夜勤直前の3連勤を許容します。"
                cards.append({
                    "id": 2,
                    "icon": "🌙",
                    "title": "夜勤直前の3連勤を許容",
                    "detail": detail,
                    "impact": "中程度",
                    "impact_color": "#fd7e14",
                    "bg": "#fffbf5",
                    "border": "#fd7e14",
                })

            # ── カード3：サブのみ ──
            if necessary[3]:
                leader_staff = [staff_names[e] for e in range(num_staff) if "主任" in str(staff_roles[e]) or "リーダー" in str(staff_roles[e])]
                sub_staff = [staff_names[e] for e in range(num_staff) if "サブ" in str(staff_roles[e])]
                # リーダーが不在になりそうな日を特定
                risky_days = []
                for d in range(num_days):
                    leader_blocked = sum(1 for e in range(num_staff)
                                        if ("主任" in str(staff_roles[e]) or "リーダー" in str(staff_roles[e]))
                                        and d in fixed_off_days_list[e])
                    total_leaders = len(leader_staff)
                    if total_leaders > 0 and leader_blocked >= total_leaders:
                        risky_days.append(f"{date_columns[d]}日({weekdays[d]})")
                leader_str = "・".join(leader_staff) if leader_staff else "リーダー"
                sub_str = "・".join(sub_staff) if sub_staff else "サブリーダー"
                if risky_days:
                    days_str = "・".join(risky_days[:3]) + ("など" if len(risky_days) > 3 else "")
                    detail = f"**{days_str}** に **{leader_str}** さんが不在になる可能性があります。その日は **{sub_str}** さん1名でまとめる構成を許容します。"
                else:
                    detail = f"**{leader_str}** さんが不在の日に、**{sub_str}** さん1名だけで日勤をまとめることを許容します。"
                cards.append({
                    "id": 3,
                    "icon": "👤",
                    "title": "役割配置をサブリーダー1名まで緩和",
                    "detail": detail,
                    "impact": "要確認",
                    "impact_color": "#dc3545",
                    "bg": "#fff8f8",
                    "border": "#dc3545",
                })

            # ── カード4：残業2連続 ──
            if necessary[4]:
                ot_staff = [staff_names[e] for e in range(num_staff) if staff_overtime_ok[e] != "×"]
                names_str = "・".join(ot_staff) if ot_staff else "残業可能スタッフ"
                # 残業が必要な日を特定
                ot_days = [f"{date_columns[d]}日({weekdays[d]})" for d in range(num_days) if overtime_req_list[d] > 0]
                if ot_days:
                    days_str = "・".join(ot_days[:4]) + ("など" if len(ot_days) > 4 else "")
                    detail = f"**{days_str}** で残業シフト(A残)が必要です。**{names_str}** さんに2日連続の残業をお願いする可能性があります。"
                else:
                    detail = f"**{names_str}** さんに、残業シフト(A残)が2日連続になる場合があります。"
                cards.append({
                    "id": 4,
                    "icon": "⏰",
                    "title": "残業(A残)の2日連続を許容",
                    "detail": detail,
                    "impact": "中程度",
                    "impact_color": "#fd7e14",
                    "bg": "#fffbf5",
                    "border": "#fd7e14",
                })

            # ── カード5：夜勤3連続 ──
            if necessary[5]:
                night_staff = [staff_names[e] for e in range(num_staff) if staff_night_ok[e] != "×"]
                names_str = "・".join(night_staff) if night_staff else "夜勤可能スタッフ"
                total_night = sum(night_req_list)
                total_cap = sum(staff_night_limits[e] for e in range(num_staff) if staff_night_ok[e] != "×")
                detail = (
                    f"月合計夜勤 **{total_night}回** 必要なのに対し、スタッフの上限合計は **{total_cap}回** です。"
                    f" **{names_str}** さんに夜勤セット(D→E→公)が月またぎで3連続になることを許容します。"
                )
                cards.append({
                    "id": 5,
                    "icon": "🌑",
                    "title": "夜勤セット3連続（月またぎ）を許容",
                    "detail": detail,
                    "impact": "要確認",
                    "impact_color": "#dc3545",
                    "bg": "#fff8f8",
                    "border": "#dc3545",
                })

            return cards

        # =============================================
        # 診断関数
        # =============================================
        def diagnose_infeasibility():
            prev_last_shift = {}
            for e, staff_name in enumerate(staff_names):
                tr = df_history[df_history.iloc[:, 0] == staff_name]
                if not tr.empty and tr.shape[1] > 5:
                    prev_last_shift[e] = str(tr.iloc[0, 5]).strip()

            forced_block = [[] for _ in range(num_days)]
            for e, staff_name in enumerate(staff_names):
                last = prev_last_shift.get(e, "")
                if last == "D":
                    if num_days > 0: forced_block[0].append((staff_name, "前月末がDのためE確定"))
                    if num_days > 1: forced_block[1].append((staff_name, "前月末がDのためE翌日→公確定"))
                elif last == "E":
                    if num_days > 0: forced_block[0].append((staff_name, "前月末がEのため公確定"))

            staff_table_rows = []
            staff_warnings = []
            for e in range(num_staff):
                total_days = num_days
                f_off = fixed_off[e]
                required_off = staff_off_days[e]
                remaining_off_to_assign = max(0, required_off - f_off)
                free_days = total_days - f_off
                work_days = free_days - remaining_off_to_assign
                chain_block = 0
                if prev_last_shift.get(e, "") == "D": chain_block = 2
                elif prev_last_shift.get(e, "") == "E": chain_block = 1
                actual_work = work_days - chain_block
                night_limit = staff_night_limits[e] if staff_night_ok[e] != "×" else 0
                night_label = f"{night_limit}回上限" if staff_night_ok[e] != "×" else "夜勤不可"
                if required_off > total_days:
                    status = "🔴 公休数が日数超過"
                    staff_warnings.append(f"**{staff_names[e]}**：公休数({required_off}日) が月の日数({total_days}日)を超えています！")
                elif actual_work < 0:
                    status = "🔴 実働日数が足りない"
                    staff_warnings.append(f"**{staff_names[e]}**：希望休({f_off}日)+残り公休({remaining_off_to_assign}日)+前月チェーン({chain_block}日)で実働できません。")
                elif actual_work <= 3:
                    status = "🟡 実働日数が少ない"
                    staff_warnings.append(f"**{staff_names[e]}**：実働可能日数が **{actual_work}日** しかありません。")
                else:
                    status = "🟢 問題なし"
                staff_table_rows.append({"スタッフ名": staff_names[e], "役割": staff_roles[e], "月日数": total_days, "希望休(固定)": f_off, "残り公休割当": remaining_off_to_assign, "前月チェーン拘束": chain_block, "実働可能日数": max(0, actual_work), "夜勤": night_label, "判定": status})

            df_staff_diag = pd.DataFrame(staff_table_rows)

            day_table_rows = []
            day_warnings = []
            for d in range(num_days):
                day_label = f"{date_columns[d]}({weekdays[d]})"
                blocked_names = []
                for e in range(num_staff):
                    if d in fixed_off_days_list[e]:
                        blocked_names.append(f"{staff_names[e]}(希望休)")
                for (sname, reason) in forced_block[d]:
                    if not any(sname in b for b in blocked_names):
                        blocked_names.append(f"{sname}({reason})")
                blocked_count = len(blocked_names)
                night_consuming = night_req_list[d]
                e_consuming = night_req_list[d - 1] if d > 0 else 0
                available_for_day = num_staff - blocked_count - night_consuming - e_consuming
                required_day = day_req_list[d]
                required_ot = overtime_req_list[d]
                total_required = required_day + required_ot
                gap = available_for_day - total_required
                if gap < 0:
                    status = f"🔴 {abs(gap)}人不足"
                    blocked_str = "、".join(blocked_names[:5]) + ("…他" if len(blocked_names) > 5 else "")
                    day_warnings.append(f"**{day_label}**：必要{total_required}人に対し最大{max(0, available_for_day)}人しか確保できません。" + (f"（{blocked_str}）" if blocked_names else ""))
                elif gap <= 1:
                    status = "🟡 ギリギリ"
                else:
                    status = "🟢 OK"
                day_table_rows.append({"日付": day_label, "必要日勤": required_day, "必要残業": required_ot, "必要夜勤(D)": night_consuming, "希望休人数": fixed_off_per_day[d], "前月チェーン拘束": len(forced_block[d]), "利用可能人数(推定)": max(0, available_for_day), "判定": status})

            df_day_diag = pd.DataFrame(day_table_rows)
            congestion_days = []
            for d in range(num_days):
                if fixed_off_per_day[d] >= max(2, num_staff // 3):
                    names_off = [staff_names[e] for e in range(num_staff) if d in fixed_off_days_list[e]]
                    congestion_days.append(f"**{date_columns[d]}日({weekdays[d]})**：{len(names_off)}名が希望休（{', '.join(names_off)}）")

            global_issues = []
            night_capable = [e for e in range(num_staff) if staff_night_ok[e] != "×"]
            total_night_capacity = sum(staff_night_limits[e] for e in night_capable)
            total_night_required = sum(night_req_list)
            if total_night_capacity < total_night_required:
                global_issues.append(f"🌙 **夜勤総量不足**：月合計夜勤 **{total_night_required}回** 必要に対し、上限合計 **{total_night_capacity}回**（不足：{total_night_required - total_night_capacity}回）")
            if len(night_capable) < 2:
                global_issues.append(f"🌙 **夜勤可能スタッフが{len(night_capable)}名**：最低2名必要です。")
            leader_count = sum(1 for r in staff_roles if "主任" in str(r) or "リーダー" in str(r) or "サブ" in str(r))
            if leader_count == 0:
                global_issues.append("👤 **リーダー/サブリーダーが0名**：毎日の日勤に必須です。")
            elif leader_count == 1:
                global_issues.append("👤 **リーダー系が1名のみ**：公休・夜勤でリーダーが不在になる日が出ます。")
            ot_capable = sum(1 for e in range(num_staff) if staff_overtime_ok[e] != "×")
            total_ot_required = sum(overtime_req_list)
            if total_ot_required > 0 and ot_capable == 0:
                global_issues.append("⏰ **残業要員が0名**：残業が必要な日がありますが「残業可否〇」のスタッフがいません。")

            return df_staff_diag, staff_warnings, df_day_diag, day_warnings, congestion_days, global_issues

        st.success("✅ データ読み込み完了。シフトを作成します。")

        # =============================================
        # ソルバー本体
        # =============================================
        def solve_shift(random_seed, allow_minus_1=False, allow_4_days=False, allow_night_3=False,
                        allow_sub_only=False, allow_ot_consec=False, allow_night_consec_3=False,
                        allow_sun_minus_1=False, allow_abs_plus_1=False):
            model = cp_model.CpModel()
            types = ['A', 'A残', 'D', 'E', '公']
            shifts = {(e, d, s): model.NewBoolVar('') for e in range(num_staff) for d in range(num_days) for s in types}

            random.seed(random_seed)
            for e in range(num_staff):
                for d in range(num_days): model.AddHint(shifts[(e, d, 'A')], random.choice([0, 1]))
                for d in range(num_days): model.AddExactlyOne(shifts[(e, d, s)] for s in types)
                if staff_night_ok[e] == "×":
                    for d in range(num_days):
                        model.Add(shifts[(e, d, 'D')] == 0); model.Add(shifts[(e, d, 'E')] == 0)
                if staff_overtime_ok[e] == "×":
                    for d in range(num_days): model.Add(shifts[(e, d, 'A残')] == 0)

            for e, staff_name in enumerate(staff_names):
                tr = df_history[df_history.iloc[:, 0] == staff_name]
                if not tr.empty:
                    last_day = str(tr.iloc[0, 5]).strip() if tr.shape[1] > 5 else ""
                    if last_day == "D":
                        model.Add(shifts[(e, 0, 'E')] == 1)
                        if num_days > 1: model.Add(shifts[(e, 1, '公')] == 1)
                    elif last_day == "E":
                        model.Add(shifts[(e, 0, '公')] == 1)

            for e in range(num_staff):
                if staff_night_ok[e] != "×":
                    tr = df_history[df_history.iloc[:, 0] == staff_names[e]]
                    if not tr.empty:
                        l_day = str(tr.iloc[0, 5]).strip() if tr.shape[1] > 5 else ""
                        if l_day != "D": model.Add(shifts[(e, 0, 'E')] == 0)
                    for d in range(num_days):
                        if d > 0: model.Add(shifts[(e, d, 'E')] == shifts[(e, d - 1, 'D')])
                        if d + 1 < num_days: model.AddImplication(shifts[(e, d, 'E')], shifts[(e, d + 1, '公')])

            penalties = []

            for e in range(num_staff):
                if staff_night_ok[e] != "×":
                    for d in range(num_days - 3): model.Add(shifts[(e, d, 'E')] + shifts[(e, d + 1, '公')] + shifts[(e, d + 2, '公')] + shifts[(e, d + 3, 'D')] <= 3)
                    for d in range(num_days - 4): model.Add(shifts[(e, d, 'E')] + shifts[(e, d + 1, '公')] + shifts[(e, d + 2, '公')] + shifts[(e, d + 3, '公')] + shifts[(e, d + 4, 'D')] <= 4)
                    tr = df_history[df_history.iloc[:, 0] == staff_names[e]]
                    if not tr.empty and tr.shape[1] > 5:
                        l_5 = [str(tr.iloc[0, i]).strip() for i in range(1, 6)]
                        if l_5[4] == "E":
                            if num_days > 2: model.Add(shifts[(e, 0, '公')] + shifts[(e, 1, '公')] + shifts[(e, 2, 'D')] <= 2)
                            if num_days > 3: model.Add(shifts[(e, 0, '公')] + shifts[(e, 1, '公')] + shifts[(e, 2, '公')] + shifts[(e, 3, 'D')] <= 3)
                        if l_5[3] == "E" and l_5[4] == "公":
                            if num_days > 1: model.Add(shifts[(e, 0, '公')] + shifts[(e, 1, 'D')] <= 1)
                            if num_days > 2: model.Add(shifts[(e, 0, '公')] + shifts[(e, 1, '公')] + shifts[(e, 2, 'D')] <= 2)

            for e, staff_name in enumerate(staff_names):
                if staff_night_ok[e] != "×":
                    past_D = [0] * 5
                    tr = df_history[df_history.iloc[:, 0] == staff_name]
                    if not tr.empty:
                        for i in range(5):
                            if (i + 1) < tr.shape[1] and str(tr.iloc[0, i + 1]).strip() == "D": past_D[i] = 1
                    all_D = past_D + [shifts[(e, d, 'D')] for d in range(num_days)]
                    for i in range(len(all_D) - 6):
                        window = all_D[i:i + 7]
                        if not allow_night_consec_3:
                            if any(isinstance(x, cp_model.IntVar) for x in window): model.Add(sum(window) <= 2)
                        else:
                            if any(isinstance(x, cp_model.IntVar) for x in window):
                                n3_var = model.NewBoolVar('')
                                model.Add(sum(window) >= 3).OnlyEnforceIf(n3_var)
                                model.Add(sum(window) <= 2).OnlyEnforceIf(n3_var.Not())
                                penalties.append(n3_var * 5000)

            for d in range(num_days):
                model.Add(sum(shifts[(e, d, 'D')] for e in range(num_staff)) == night_req_list[d])
                model.Add(sum(shifts[(e, d, 'A残')] for e in range(num_staff)) == overtime_req_list[d])
                act_day = sum((shifts[(e, d, 'A')] + shifts[(e, d, 'A残')]) for e in range(num_staff) if "新人" not in str(staff_roles[e]))
                req = day_req_list[d]
                is_sun = ('日' in weekdays[d])
                is_abs = (absolute_req_list[d] == "〇")
                if is_abs:
                    model.Add(act_day >= req)
                    if not allow_abs_plus_1:
                        over_var = model.NewIntVar(0, 100, ''); diff = model.NewIntVar(-100, 100, '')
                        model.Add(diff == act_day - req); model.AddMaxEquality(over_var, [0, diff])
                        penalties.append(over_var * 1)
                    else:
                        over_var = model.NewIntVar(0, 100, ''); diff = model.NewIntVar(-100, 100, '')
                        model.Add(diff == act_day - (req + 1)); model.AddMaxEquality(over_var, [0, diff])
                        penalties.append(over_var * 500)
                elif is_sun:
                    model.Add(act_day <= req)
                    if not allow_sun_minus_1:
                        model.Add(act_day == req)
                    else:
                        leader_present = sum((1 if "主任" in str(staff_roles[e]) or "リーダー" in str(staff_roles[e]) else 0) * (shifts[(e, d, 'A')] + shifts[(e, d, 'A残')]) for e in range(num_staff))
                        has_leader = model.NewBoolVar('')
                        model.Add(leader_present >= 1).OnlyEnforceIf(has_leader)
                        model.Add(leader_present == 0).OnlyEnforceIf(has_leader.Not())
                        model.Add(act_day >= req - 1).OnlyEnforceIf(has_leader)
                        model.Add(act_day >= req).OnlyEnforceIf(has_leader.Not())
                        m_var = model.NewBoolVar('')
                        model.Add(act_day == req - 1).OnlyEnforceIf(m_var)
                        model.Add(act_day != req - 1).OnlyEnforceIf(m_var.Not())
                        penalties.append(m_var * 1000)
                else:
                    if not allow_minus_1: model.Add(act_day >= req)
                    else:
                        model.Add(act_day >= req - 1); m_var = model.NewBoolVar('')
                        model.Add(act_day == req - 1).OnlyEnforceIf(m_var); model.Add(act_day != req - 1).OnlyEnforceIf(m_var.Not())
                        penalties.append(m_var * 1000)
                    over_var = model.NewIntVar(0, 100, ''); diff = model.NewIntVar(-100, 100, '')
                    model.Add(diff == act_day - req); model.AddMaxEquality(over_var, [0, diff])
                    penalties.append(over_var * 100)

                l_score = sum((2 if "主任" in str(staff_roles[e]) or "リーダー" in str(staff_roles[e]) else 1 if "サブ" in str(staff_roles[e]) else 0) * (shifts[(e, d, 'A')] + shifts[(e, d, 'A残')]) for e in range(num_staff))
                if not allow_sub_only: model.Add(l_score >= 2)
                else:
                    model.Add(l_score >= 1); sub_var = model.NewBoolVar('')
                    model.Add(l_score == 1).OnlyEnforceIf(sub_var); penalties.append(sub_var * 1000)

            for e, staff_name in enumerate(staff_names):
                tr = df_history[df_history.iloc[:, 0] == staff_name]
                if not tr.empty:
                    history_today_cols = list(tr.columns[6:])
                    for d in range(num_days):
                        col_name = date_columns[d]
                        if col_name in history_today_cols and str(tr.iloc[0][col_name]).strip() == "公":
                            model.Add(shifts[(e, d, '公')] == 1)

            for e in range(num_staff):
                model.Add(sum(shifts[(e, d, '公')] for d in range(num_days)) == int(staff_off_days[e]))
                if staff_night_ok[e] != "×": model.Add(sum(shifts[(e, d, 'D')] for d in range(num_days)) <= int(staff_night_limits[e]))

            limit_groups = {}
            for e in range(num_staff):
                if staff_night_ok[e] != "×" and int(staff_night_limits[e]) > 0:
                    limit_groups.setdefault(int(staff_night_limits[e]), []).append(e)
            for limit, members in limit_groups.items():
                if len(members) >= 2:
                    actual_nights = [sum(shifts[(m, d, 'D')] for d in range(num_days)) for m in members]
                    max_n = model.NewIntVar(0, limit, ''); min_n = model.NewIntVar(0, limit, '')
                    model.AddMaxEquality(max_n, actual_nights); model.AddMinEquality(min_n, actual_nights)
                    model.Add(max_n - min_n <= 1)

            for e in range(num_staff):
                for d in range(num_days - 3): model.Add(shifts[(e, d, '公')] + shifts[(e, d + 1, '公')] + shifts[(e, d + 2, '公')] + shifts[(e, d + 3, '公')] <= 3)
                for d in range(num_days - 2):
                    is_3_off = model.NewBoolVar('')
                    model.Add(shifts[(e, d, '公')] + shifts[(e, d + 1, '公')] + shifts[(e, d + 2, '公')] == 3).OnlyEnforceIf(is_3_off)
                    model.Add(shifts[(e, d, '公')] + shifts[(e, d + 1, '公')] + shifts[(e, d + 2, '公')] <= 2).OnlyEnforceIf(is_3_off.Not())
                    penalties.append(is_3_off * 500)
                is_2_offs = []
                for d in range(num_days - 1):
                    is_2_off = model.NewBoolVar('')
                    model.Add(shifts[(e, d, '公')] + shifts[(e, d + 1, '公')] == 2).OnlyEnforceIf(is_2_off)
                    model.Add(shifts[(e, d, '公')] + shifts[(e, d + 1, '公')] <= 1).OnlyEnforceIf(is_2_off.Not())
                    is_2_offs.append(is_2_off)
                has_any_2_off = model.NewBoolVar('')
                model.Add(sum(is_2_offs) >= 1).OnlyEnforceIf(has_any_2_off); model.Add(sum(is_2_offs) == 0).OnlyEnforceIf(has_any_2_off.Not())
                penalties.append(has_any_2_off.Not() * 300)

            for e in range(num_staff):
                target_lvl = staff_comp_lvl[e]
                w_base = 10 ** target_lvl if target_lvl > 0 else 0
                for d in range(num_days - 3):
                    def work(day): return shifts[(e, day, 'A')] + shifts[(e, day, 'A残')]
                    if allow_4_days and target_lvl > 0:
                        if d < num_days - 4: model.Add(work(d) + work(d + 1) + work(d + 2) + work(d + 3) + work(d + 4) <= 4)
                        p_var = model.NewBoolVar('')
                        model.Add(work(d) + work(d + 1) + work(d + 2) + work(d + 3) == 4).OnlyEnforceIf(p_var)
                        model.Add(work(d) + work(d + 1) + work(d + 2) + work(d + 3) <= 3).OnlyEnforceIf(p_var.Not())
                        penalties.append(p_var * w_base)
                    else:
                        model.Add(work(d) + work(d + 1) + work(d + 2) + work(d + 3) <= 3)

                    if allow_night_3 and target_lvl > 0:
                        np_var = model.NewBoolVar('')
                        model.Add(work(d) + work(d + 1) + work(d + 2) == 3).OnlyEnforceIf(np_var)
                        model.Add(work(d) + work(d + 1) + work(d + 2) <= 2).OnlyEnforceIf(np_var.Not())
                        final_p = model.NewIntVar(0, w_base, '')
                        model.AddMultiplicationEquality(final_p, [np_var, shifts[(e, d + 3, 'D')]])
                        penalties.append(final_p)
                    else:
                        model.Add(work(d) + work(d + 1) + work(d + 2) <= 2).OnlyEnforceIf(shifts[(e, d + 3, 'D')])

            for e in range(num_staff):
                for d in range(num_days - 1):
                    if not allow_ot_consec: model.Add(shifts[(e, d, 'A残')] + shifts[(e, d + 1, 'A残')] <= 1)
                    else:
                        ot_var = model.NewBoolVar('')
                        model.Add(shifts[(e, d, 'A残')] + shifts[(e, d + 1, 'A残')] == 2).OnlyEnforceIf(ot_var)
                        penalties.append(ot_var * 500)

            for e in range(num_staff):
                if staff_overtime_ok[e] != "×":
                    total_day_work = sum(shifts[(e, d, 'A')] + shifts[(e, d, 'A残')] for d in range(num_days))
                    b_has_work = model.NewBoolVar('')
                    model.Add(total_day_work > 0).OnlyEnforceIf(b_has_work); model.Add(total_day_work == 0).OnlyEnforceIf(b_has_work.Not())
                    min_a = int(staff_min_normal_a[e])
                    total_a_normal = sum(shifts[(e, d, 'A')] for d in range(num_days))
                    model.Add(total_a_normal >= min_a).OnlyEnforceIf(b_has_work)

            ot_burden_scores = []
            for e in range(num_staff):
                if staff_overtime_ok[e] != "×":
                    total_work_score = sum(shifts[(e, d, 'A')] + (shifts[(e, d, 'A残')] * 2) for d in range(num_days))
                    ot_burden_scores.append(total_work_score)
            if ot_burden_scores:
                max_b = model.NewIntVar(0, 100, ''); min_b = model.NewIntVar(0, 100, '')
                model.AddMaxEquality(max_b, ot_burden_scores); model.AddMinEquality(min_b, ot_burden_scores)
                penalties.append((max_b - min_b) * 50)

            for e in range(num_staff):
                ot_bias = random.randint(-2, 2); night_bias = random.randint(-2, 2); off_bias = random.randint(-2, 2)
                if staff_overtime_ok[e] != "×": penalties.append(sum(shifts[(e, d, 'A残')] for d in range(num_days)) * ot_bias)
                if staff_night_ok[e] != "×": penalties.append(sum(shifts[(e, d, 'D')] for d in range(num_days)) * night_bias)
                penalties.append(sum(shifts[(e, d, '公')] for d in range(num_days)) * off_bias)
                for d in range(num_days): penalties.append(shifts[(e, d, 'A')] * random.randint(-1, 1))

            if penalties: model.Minimize(sum(penalties))

            solver = cp_model.CpSolver()
            solver.parameters.max_time_in_seconds = 45.0
            solver.parameters.random_seed = random_seed
            status = solver.Solve(model)

            if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE: return solver, shifts
            else: return None, None

        def solve_shift_fast(flags):
            from ortools.sat.python import cp_model as _cp
            m1, f4, n3, sub, ot, nc3, sun1 = flags
            abs1 = True
            _model = _cp.CpModel()
            _types = ['A', 'A残', 'D', 'E', '公']
            _shifts = {(e, d, s): _model.NewBoolVar('') for e in range(num_staff) for d in range(num_days) for s in _types}
            for e in range(num_staff):
                for d in range(num_days): _model.AddExactlyOne(_shifts[(e, d, s)] for s in _types)
                if staff_night_ok[e] == "×":
                    for d in range(num_days): _model.Add(_shifts[(e, d, 'D')] == 0); _model.Add(_shifts[(e, d, 'E')] == 0)
                if staff_overtime_ok[e] == "×":
                    for d in range(num_days): _model.Add(_shifts[(e, d, 'A残')] == 0)
            for e, staff_name in enumerate(staff_names):
                tr = df_history[df_history.iloc[:, 0] == staff_name]
                if not tr.empty:
                    last_day = str(tr.iloc[0, 5]).strip() if tr.shape[1] > 5 else ""
                    if last_day == "D":
                        _model.Add(_shifts[(e, 0, 'E')] == 1)
                        if num_days > 1: _model.Add(_shifts[(e, 1, '公')] == 1)
                    elif last_day == "E": _model.Add(_shifts[(e, 0, '公')] == 1)
            for e in range(num_staff):
                if staff_night_ok[e] != "×":
                    tr = df_history[df_history.iloc[:, 0] == staff_names[e]]
                    if not tr.empty:
                        l_day = str(tr.iloc[0, 5]).strip() if tr.shape[1] > 5 else ""
                        if l_day != "D": _model.Add(_shifts[(e, 0, 'E')] == 0)
                    for d in range(num_days):
                        if d > 0: _model.Add(_shifts[(e, d, 'E')] == _shifts[(e, d - 1, 'D')])
                        if d + 1 < num_days: _model.AddImplication(_shifts[(e, d, 'E')], _shifts[(e, d + 1, '公')])
            for e, staff_name in enumerate(staff_names):
                tr = df_history[df_history.iloc[:, 0] == staff_name]
                if not tr.empty:
                    history_today_cols = list(tr.columns[6:])
                    for d in range(num_days):
                        col_name = date_columns[d]
                        if col_name in history_today_cols and str(tr.iloc[0][col_name]).strip() == "公":
                            _model.Add(_shifts[(e, d, '公')] == 1)
            for e in range(num_staff):
                _model.Add(sum(_shifts[(e, d, '公')] for d in range(num_days)) == int(staff_off_days[e]))
                if staff_night_ok[e] != "×": _model.Add(sum(_shifts[(e, d, 'D')] for d in range(num_days)) <= int(staff_night_limits[e]))
            for e in range(num_staff):
                if staff_night_ok[e] != "×":
                    past_D = [0] * 5
                    tr = df_history[df_history.iloc[:, 0] == staff_names[e]]
                    if not tr.empty:
                        for i in range(5):
                            if (i + 1) < tr.shape[1] and str(tr.iloc[0, i + 1]).strip() == "D": past_D[i] = 1
                    all_D = past_D + [_shifts[(e, d, 'D')] for d in range(num_days)]
                    for i in range(len(all_D) - 6):
                        window = all_D[i:i + 7]
                        if not nc3 and any(isinstance(x, _cp.IntVar) for x in window): _model.Add(sum(window) <= 2)
            for d in range(num_days):
                _model.Add(sum(_shifts[(e, d, 'D')] for e in range(num_staff)) == night_req_list[d])
                _model.Add(sum(_shifts[(e, d, 'A残')] for e in range(num_staff)) == overtime_req_list[d])
                act_day = sum((_shifts[(e, d, 'A')] + _shifts[(e, d, 'A残')]) for e in range(num_staff) if "新人" not in str(staff_roles[e]))
                req = day_req_list[d]; is_sun = ('日' in weekdays[d]); is_abs = (absolute_req_list[d] == "〇")
                if is_abs:
                    _model.Add(act_day >= req)
                    if abs1: _model.Add(act_day <= req + 1)
                elif is_sun:
                    _model.Add(act_day <= req)
                    if not sun1: _model.Add(act_day == req)
                    else: _model.Add(act_day >= req - 1)
                else:
                    if not m1: _model.Add(act_day >= req)
                    else: _model.Add(act_day >= req - 1)
                l_score = sum((2 if "主任" in str(staff_roles[e]) or "リーダー" in str(staff_roles[e]) else 1 if "サブ" in str(staff_roles[e]) else 0) * (_shifts[(e, d, 'A')] + _shifts[(e, d, 'A残')]) for e in range(num_staff))
                if not sub: _model.Add(l_score >= 2)
                else: _model.Add(l_score >= 1)
            for e in range(num_staff):
                target_lvl = staff_comp_lvl[e]
                for d in range(num_days - 3):
                    def _work(day, _e=e): return _shifts[(_e, day, 'A')] + _shifts[(_e, day, 'A残')]
                    if not f4 or target_lvl == 0: _model.Add(_work(d) + _work(d + 1) + _work(d + 2) + _work(d + 3) <= 3)
                    if not n3 or target_lvl == 0: _model.Add(_work(d) + _work(d + 1) + _work(d + 2) <= 2).OnlyEnforceIf(_shifts[(e, d + 3, 'D')])
            for e in range(num_staff):
                for d in range(num_days - 1):
                    if not ot: _model.Add(_shifts[(e, d, 'A残')] + _shifts[(e, d + 1, 'A残')] <= 1)
            for e in range(num_staff):
                if staff_overtime_ok[e] != "×":
                    total_day_work = sum(_shifts[(e, d, 'A')] + _shifts[(e, d, 'A残')] for d in range(num_days))
                    b_has = _model.NewBoolVar('')
                    _model.Add(total_day_work > 0).OnlyEnforceIf(b_has); _model.Add(total_day_work == 0).OnlyEnforceIf(b_has.Not())
                    _model.Add(sum(_shifts[(e, d, 'A')] for d in range(num_days)) >= int(staff_min_normal_a[e])).OnlyEnforceIf(b_has)
            _solver = _cp.CpSolver()
            _solver.parameters.max_time_in_seconds = 12.0
            _solver.parameters.random_seed = 1
            _status = _solver.Solve(_model)
            return _status in (_cp.OPTIMAL, _cp.FEASIBLE)

        COMPROMISE_LABELS = [
            "平日・祝の日勤人数を-1にする",
            "最大4連勤のお願い",
            "夜勤前3日連続日勤のお願い",
            "役割配置をサブ1名まで下げる",
            "残業(A残)の2日連続",
            "夜勤セット3連続（月またぎ含む）",
            "日曜の出勤人数を-1にする（リーダー在勤条件）",
        ]
        ALL_ON = (True, True, True, True, True, True, True)

        if 'min_compromise_result' not in st.session_state:
            st.session_state.min_compromise_result = None
        if 'step3_failed' not in st.session_state:
            st.session_state.step3_failed = False
        if 'card_selections' not in st.session_state:
            st.session_state.card_selections = {}

        ALL_SEEDS = [7, 42, 137, 512, 9999, 31415, 271828, 100003, 777777, 999983]
        num_patterns = st.selectbox("🔢 作成するシフトのパターン数", [1, 2, 3, 4, 5], index=2)
        use_seeds = ALL_SEEDS[:num_patterns]

        if not st.session_state.needs_compromise:
            if st.button(f"▶️ 【STEP 1】まずは妥協なしで理想のシフトを計算する（{num_patterns}パターン）"):
                with st.spinner(f'AIが「妥協なし」の完璧なシフトを{num_patterns}パターン模索中...'):
                    results = []
                    for seed in use_seeds:
                        solver, shifts = solve_shift(seed, False, False, False, False, False, False, False, True)
                        if solver: results.append((solver, shifts))
                    if results:
                        st.success(f"🎉 妥協なしで完璧なシフトが {len(results)} パターン組めました！")
                    else:
                        st.session_state.needs_compromise = True
                        st.session_state.min_compromise_result = None
                        st.session_state.step3_failed = False
                        st.rerun()

        else:
            st.error("⚠️ 妥協なしではシフトを組めませんでした。下の提案を確認してください。")

            # ── 診断レポート ──
            df_staff_diag, staff_warnings, df_day_diag, day_warnings, congestion_days, global_issues = diagnose_infeasibility()
            with st.expander("🔍 **【原因診断レポート】** ← クリックして確認", expanded=False):
                if global_issues:
                    st.markdown("### ⚠️ 全体チェックで検出された問題")
                    for g in global_issues: st.error(g)
                    st.markdown("---")
                st.markdown("### 👥 ① スタッフ別 稼働余力チェック")
                def color_staff_status(val):
                    if "🔴" in str(val): return "background-color:#FFD0D0;font-weight:bold;"
                    if "🟡" in str(val): return "background-color:#FFF5CC;"
                    if "🟢" in str(val): return "background-color:#D8F5D8;"
                    return ""
                st.dataframe(df_staff_diag.style.applymap(color_staff_status, subset=["判定"]), use_container_width=True)
                if staff_warnings:
                    for w in staff_warnings: st.warning(w)
                st.markdown("---")
                st.markdown("### 📅 ② 日別 人員過不足チェック")
                def color_day_status(val):
                    if "🔴" in str(val): return "background-color:#FFD0D0;font-weight:bold;"
                    if "🟡" in str(val): return "background-color:#FFF5CC;"
                    if "🟢" in str(val): return "background-color:#D8F5D8;"
                    return ""
                st.dataframe(df_day_diag.style.applymap(color_day_status, subset=["判定"]), use_container_width=True)
                if day_warnings:
                    for w in day_warnings: st.error(w)
                if congestion_days:
                    st.markdown("### 🗓️ ③ 希望休が集中している日")
                    for c in congestion_days: st.warning(c)
                st.caption("※ この診断は推定値です。")

            st.markdown("---")

            # ── STEP2：自動分析ボタン ──
            if st.button("🔎 【STEP 2】どの妥協が必要か自動で調べる（約1分）"):
                progress = st.progress(0, text="全妥協案ONで確認中...")
                can_all = solve_shift_fast(ALL_ON)
                if not can_all:
                    st.error("😭 全妥協案をONにしても組めません。希望休や人数設定を見直してください。")
                    st.session_state.min_compromise_result = None
                else:
                    necessary = [True] * 7
                    for i in range(7):
                        progress.progress((i + 1) / 8, text=f"「{COMPROMISE_LABELS[i]}」が不要か確認中... ({i + 1}/7)")
                        flags_test = list(ALL_ON)
                        flags_test[i] = False
                        if solve_shift_fast(tuple(flags_test)):
                            necessary[i] = False
                    progress.progress(1.0, text="分析完了！")
                    st.session_state.min_compromise_result = necessary
                    # カード選択を初期化（必要なカードはデフォルトON）
                    st.session_state.card_selections = {i: necessary[i] for i in range(7)}

            # =============================================
            # 🃏 AI交渉カード UI（★ 新機能）
            # =============================================
            if st.session_state.min_compromise_result is not None:
                necessary = st.session_state.min_compromise_result
                cards = generate_compromise_cards(necessary)

                needed_count = sum(1 for c in cards)
                unneeded_labels = [COMPROMISE_LABELS[i] for i in range(6) if not necessary[i]]

                st.markdown("---")
                st.markdown("### 🃏 【STEP 3】AIからの提案カード")
                st.markdown(f"以下 **{needed_count}件** の妥協が必要です。各カードを確認して「許可する」をON/OFFしてください。")

                if unneeded_labels:
                    with st.expander(f"✅ 不要と判定された妥協案（{len(unneeded_labels)}件）"):
                        for u in unneeded_labels:
                            st.markdown(f"　🟢 ~~{u}~~")

                st.markdown("")

                # カードを2列で表示
                card_cols = st.columns(2)
                for idx, card in enumerate(cards):
                    col = card_cols[idx % 2]
                    cid = card["id"]
                    with col:
                        # カードのHTML風表示
                        st.markdown(
                            f"""
                            <div style="
                                background: {card['bg']};
                                border: 2px solid {card['border']};
                                border-radius: 12px;
                                padding: 16px 18px;
                                margin-bottom: 8px;
                            ">
                                <div style="font-size:1.3em; font-weight:bold;">{card['icon']} {card['title']}</div>
                                <div style="margin: 8px 0; font-size:0.85em; color:#555;">
                                    影響度：<span style="color:{card['impact_color']}; font-weight:bold;">{card['impact']}</span>
                                </div>
                                <div style="font-size:0.9em; line-height:1.6;">{card['detail']}</div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )
                        current_val = st.session_state.card_selections.get(cid, True)
                        new_val = st.checkbox(
                            "✅ この妥協を許可する",
                            value=current_val,
                            key=f"card_check_{cid}"
                        )
                        st.session_state.card_selections[cid] = new_val
                        st.markdown("")

                # 選択結果をフラグに変換
                sel = st.session_state.card_selections
                allow_minus_1         = sel.get(0, False)
                allow_4_days          = sel.get(1, False)
                allow_night_3         = sel.get(2, False)
                allow_sub_only        = sel.get(3, False)
                allow_ot_consec       = sel.get(4, False)
                allow_night_consec_3  = sel.get(5, False)
                allow_abs_plus_1      = True

                # 許可したカードのサマリー
                accepted = [cards[i]["title"] for i in range(len(cards)) if sel.get(cards[i]["id"], False)]
                rejected = [cards[i]["title"] for i in range(len(cards)) if not sel.get(cards[i]["id"], False)]

                c1, c2 = st.columns(2)
                with c1:
                    if accepted:
                        st.success("**許可した妥協案**\n\n" + "\n".join(f"✅ {a}" for a in accepted))
                with c2:
                    if rejected:
                        st.error("**却下した妥協案**\n\n" + "\n".join(f"❌ {r}" for r in rejected))

                st.markdown("---")

                if st.button(f"🔄 【STEP 3】選んだ妥協案で{num_patterns}パターン作成"):
                    with st.spinner('計算中...'):
                        results = []
                        for seed in use_seeds:
                            solver, shifts = solve_shift(seed, allow_minus_1, allow_4_days, allow_night_3, allow_sub_only, allow_ot_consec, allow_night_consec_3, False, allow_abs_plus_1)
                            if solver: results.append((solver, shifts))
                        if results:
                            st.success(f"✨ {len(results)}パターン完成！")
                            st.session_state.needs_compromise = False
                            st.session_state.step3_failed = False
                        else:
                            st.session_state.step3_failed = True
                            st.rerun()

                # ── STEP4：日曜-1 ──
                if st.session_state.step3_failed:
                    st.error("❌ 選んだ妥協案でもシフトが組めませんでした。")
                    st.markdown("---")
                    st.markdown("### 🌟 【STEP 4】追加の切り札カード")

                    sun_needed = necessary[6] if len(necessary) > 6 else False
                    sun_risky = []
                    for d in range(num_days):
                        if '日' in weekdays[d]:
                            blocked = sum(1 for e in range(num_staff) if d in fixed_off_days_list[e])
                            available = num_staff - blocked - night_req_list[d]
                            if available < day_req_list[d]:
                                sun_risky.append(f"{date_columns[d]}日(日)")

                    sun_detail = ""
                    if sun_risky:
                        days_str = "・".join(sun_risky[:3]) + ("など" if len(sun_risky) > 3 else "")
                        sun_detail = f"**{days_str}** の日曜日で日勤人数が不足する可能性があります。リーダーが在勤している日曜に限り、設定人数より1名少ない状態を許容します。"
                    else:
                        sun_detail = "一部の日曜日で、リーダーが在勤している場合に限り設定人数より1名少ない状態を許容します。"

                    st.markdown(
                        f"""
                        <div style="
                            background: #fff0f0;
                            border: 2px solid #c0392b;
                            border-radius: 12px;
                            padding: 16px 18px;
                            margin-bottom: 12px;
                        ">
                            <div style="font-size:1.3em; font-weight:bold;">🌅 日曜の出勤人数を-1にする</div>
                            <div style="margin: 8px 0; font-size:0.85em; color:#555;">
                                影響度：<span style="color:#c0392b; font-weight:bold;">最終手段</span>
                            </div>
                            <div style="font-size:0.9em; line-height:1.6;">{sun_detail}</div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
                    allow_sun_minus_1 = st.checkbox("✅ 日曜-1を許可する（リーダー在勤時のみ）", value=sun_needed, key="card_sun")

                    if st.button(f"🔄 【STEP 4】日曜-1も加えて{num_patterns}パターン作成"):
                        with st.spinner('計算中...'):
                            results = []
                            for seed in use_seeds:
                                solver, shifts = solve_shift(seed, allow_minus_1, allow_4_days, allow_night_3, allow_sub_only, allow_ot_consec, allow_night_consec_3, allow_sun_minus_1, allow_abs_plus_1)
                                if solver: results.append((solver, shifts))
                            if results:
                                st.success(f"✨ {len(results)}パターン完成！")
                                st.session_state.needs_compromise = False
                                st.session_state.step3_failed = False
                            else:
                                st.error("😭 まだ組めません。希望休や人数設定を見直してください。")

        # =============================================
        # 結果表示・Excel出力（変更なし）
        # =============================================
        if 'results' in locals() and results:
            cols = []
            for d_val, w_val in zip(date_columns, weekdays):
                try:
                    dt = datetime.date(target_year, target_month, int(d_val))
                    if jpholiday.is_holiday(dt): cols.append(f"{d_val}({w_val}・祝)")
                    else: cols.append(f"{d_val}({w_val})")
                except ValueError: cols.append(f"{d_val}({w_val})")

            hope_off_set = set()       # 希望休（黄色）
            hope_shift_dict = {}       # 希望シフト（水色）: (e, d) -> シフト種別
            VALID_SHIFTS = {'A', 'A残', 'D', 'E', '公'}
            for e, staff_name in enumerate(staff_names):
                tr = df_history[df_history.iloc[:, 0] == staff_name]
                if not tr.empty:
                    history_today_cols = list(tr.columns[6:])
                    for d in range(num_days):
                        col_name = date_columns[d]
                        if col_name in history_today_cols:
                            val = str(tr.iloc[0][col_name]).strip()
                            if val == "公":
                                hope_off_set.add((e, d))
                            elif val in VALID_SHIFTS:
                                hope_shift_dict[(e, d)] = val

            tabs = st.tabs([f"提案パターン {i + 1}" for i in range(len(results))])
            for i, (solver, shifts) in enumerate(results):
                with tabs[i]:
                    data = []
                    for e in range(num_staff):
                        row = {"スタッフ名": staff_names[e]}
                        for d in range(num_days):
                            for s in ['A', 'A残', 'D', 'E', '公']:
                                if solver.Value(shifts[(e, d, s)]):
                                    if (s == 'A' or s == 'A残') and str(staff_part_shifts[e]).strip() not in ["", "nan"]: row[cols[d]] = str(staff_part_shifts[e]).strip()
                                    else: row[cols[d]] = s
                        data.append(row)

                    df_res = pd.DataFrame(data)
                    df_res['日勤(A/P)回数'] = df_res[cols].apply(lambda x: x.str.contains('A|P|Ｐ', na=False) & ~x.str.contains('残', na=False)).sum(axis=1)
                    df_res['残業(A残)回数'] = (df_res[cols] == 'A残').sum(axis=1)
                    df_res['夜勤(D)回数'] = (df_res[cols] == 'D').sum(axis=1)
                    df_res['公休回数'] = (df_res[cols] == '公').sum(axis=1)

                    sum_A = {"スタッフ名": "【日勤(A/P) 合計人数】"}
                    sum_Az = {"スタッフ名": "【残業(A残) 合計人数】"}
                    sum_D = {"スタッフ名": "【夜勤(D) 人数】"}
                    sum_E = {"スタッフ名": "【夜勤明け(E) 人数】"}
                    for c in ['日勤(A/P)回数', '残業(A残)回数', '夜勤(D)回数', '公休回数']:
                        sum_A[c] = ""; sum_Az[c] = ""; sum_D[c] = ""; sum_E[c] = ""
                    for d, c in enumerate(cols):
                        sum_A[c] = sum(1 for e in range(num_staff) if str(df_res.loc[e, c]) in ['A', 'A残'] or 'P' in str(df_res.loc[e, c]) and "新人" not in str(staff_roles[e]))
                        sum_Az[c] = (df_res[c] == 'A残').sum()
                        sum_D[c] = (df_res[c] == 'D').sum()
                        sum_E[c] = (df_res[c] == 'E').sum()

                    df_fin = pd.concat([df_res, pd.DataFrame([sum_A, sum_Az, sum_D, sum_E])], ignore_index=True)

                    def highlight_warnings(df):
                        styles = pd.DataFrame('', index=df.index, columns=df.columns)
                        for d, col_name in enumerate(cols):
                            if "土" in col_name: styles.iloc[:, d + 1] = 'background-color: #E6F2FF;'
                            elif "日" in col_name or "祝" in col_name: styles.iloc[:, d + 1] = 'background-color: #FFE6E6;'
                        for e in range(num_staff):
                            for d in range(num_days):
                                if (e, d) in hope_off_set:
                                    styles.loc[e, cols[d]] = 'background-color: #FFFF00; font-weight: bold;'
                                elif (e, d) in hope_shift_dict:
                                    styles.loc[e, cols[d]] = 'background-color: #FFFF00; font-weight: bold;'
                        for d, col_name in enumerate(cols):
                            actual_a = df.loc[len(staff_names), col_name]
                            target_a = day_req_list[d]
                            if actual_a != "":
                                if actual_a < target_a: styles.loc[len(staff_names), col_name] = 'background-color: #FFCCCC; color: red; font-weight: bold;'
                                elif actual_a > target_a: styles.loc[len(staff_names), col_name] = 'background-color: #CCFFFF; color: blue; font-weight: bold;'
                        for e in range(num_staff):
                            for d in range(num_days):
                                def is_day_work(day_idx):
                                    if day_idx >= num_days: return False
                                    v = str(df.loc[e, cols[day_idx]])
                                    return v == 'A' or v == 'A残' or 'P' in v or 'Ｐ' in v
                                if is_day_work(d) and is_day_work(d + 1) and is_day_work(d + 2) and is_day_work(d + 3):
                                    for k in range(4): styles.loc[e, cols[d + k]] = 'background-color: #FFFF99; font-weight: bold; color: black;'
                                if d + 3 < num_days:
                                    if is_day_work(d) and is_day_work(d + 1) and is_day_work(d + 2) and str(df.loc[e, cols[d + 3]]) == 'D':
                                        for k in range(4): styles.loc[e, cols[d + k]] = 'background-color: #FFD580; font-weight: bold; color: black;'
                                if d + 8 < num_days:
                                    if str(df.loc[e, cols[d]]) == 'D' and str(df.loc[e, cols[d + 3]]) == 'D' and str(df.loc[e, cols[d + 6]]) == 'D':
                                        for k in range(9): styles.loc[e, cols[d + k]] = 'background-color: #E6E6FA; font-weight: bold; color: black;'
                        return styles

                    st.dataframe(df_fin.style.apply(highlight_warnings, axis=None))

                    # =============================================
                    # 📊 振り返りレポート & 有休・夏冬休管理
                    # =============================================
                    with st.expander("📊 振り返りレポート & 有休・休暇管理を見る", expanded=False):
                        df_report, alerts, red_alerts, march_forced = build_review_report(df_fin, cols)

                        # ── 休暇優先順位の説明 ──
                        st.info("📋 **休暇優先順位**：公休（9日）＞ 夏季休暇（7〜9月・3日）＞ 冬季休暇（12〜2月・4日）＞ 年休（4〜2月・年5日）")
                        st.markdown("---")

                        # ── 🚨 赤色：強制対応アラート ──
                        if red_alerts:
                            st.markdown("### 🚨 要対応（今月シフトに必ず組み込んでください）")
                            for a in red_alerts:
                                st.error(a)
                            st.markdown("---")

                        # ── 3月強制消化スタッフ一覧 ──
                        if target_month == 3 and march_forced:
                            st.error("### 🚨 3月シフト：以下のスタッフに有休を必ず組み込んでください")
                            for mf in march_forced:
                                st.error(f"　👤 **{mf['name']}**：有休 **{mf['days']}日** を3月中に取得させてください。")
                            st.markdown("---")

                        # ── 🟡 黄色：注意アラート ──
                        if alerts:
                            st.markdown("### ⚠️ 注意・促しアラート")
                            for a in alerts:
                                st.warning(a)
                            st.markdown("---")

                        # ── スタッフ別サマリー表 ──
                        st.markdown("### 👥 スタッフ別サマリー")
                        display_cols = [
                            "スタッフ名", "役割",
                            "今月公休", "今月夜勤", "今月残業", "今月最大連勤",
                            "夜勤累計(年)", "残業累計(年)",
                            "有休取得済", "有休残", "年間義務残",
                            "夏休残", "冬休残", "判定"
                        ]
                        df_disp = df_report[display_cols].copy()

                        def color_report(val):
                            s = str(val)
                            if "🔴" in s or "🚨" in s: return "background-color:#FFD0D0; font-weight:bold;"
                            if "🟡" in s or "⚠️" in s: return "background-color:#FFF5CC;"
                            if "🎉" in s: return "background-color:#D8F5D8; font-weight:bold;"
                            if "🟢" in s: return "background-color:#D8F5D8;"
                            return ""

                        st.dataframe(
                            df_disp.style.applymap(color_report, subset=["判定"]),
                            use_container_width=True
                        )

                        # ── 有休進捗バー ──
                        st.markdown("### 📅 有休取得進捗（4月〜2月 年5日ルール）")
                        remain_m = remaining_months_to_feb(target_month)
                        if target_month == 3:
                            st.error("📌 3月は強制消化月です。未達スタッフは必ずシフトに有休を組み込んでください。")
                        elif remain_m == 0:
                            st.info("📌 2月末が締め切りです。今月取得できなかった分は3月に強制消化になります。")
                        else:
                            st.info(f"📌 現在 **{target_month}月**。2月末まで残り **{remain_m}ヶ月**。年5日取得が目標です（優先度：公休＞夏冬休＞年休）。")

                        progress_data = df_report[["スタッフ名", "有休取得済", "年間義務残"]].set_index("スタッフ名")
                        st.bar_chart(progress_data)

                        # ── 夏冬休グラフ（対象期間のみ） ──
                        if target_month in SUMMER_MONTHS:
                            st.markdown("### ☀️ 夏季休暇残日数（7〜9月で3日取得）")
                            st.bar_chart(df_report[["スタッフ名", "夏休残"]].set_index("スタッフ名"))
                        if target_month in WINTER_MONTHS:
                            st.markdown("### ❄️ 冬季休暇残日数（12〜2月で4日取得）")
                            st.bar_chart(df_report[["スタッフ名", "冬休残"]].set_index("スタッフ名"))

                        # ── 夜勤・残業グラフ ──
                        st.markdown("### 📈 今月の夜勤・残業回数")
                        st.bar_chart(df_report[["スタッフ名", "今月夜勤", "今月残業"]].set_index("スタッフ名"))

                        st.markdown("---")
                        st.info("💾 年間管理シートは下の「📥 シフト管理ファイルをダウンロード」に含まれています。")

                    # =============================================
                    # 📥 統合Excelダウンロード（1ファイルに全シート）
                    # =============================================
                    st.markdown("---")
                    st.markdown(f"### 📥 パターン {i + 1} をダウンロード")
                    st.caption("完成シフト・予定実績・年間管理がすべて1ファイルに含まれます。来月もこのファイルをアップロードしてください。")

                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:

                        # ── ① 完成シフトシート ──
                        df_fin.to_excel(writer, index=False, sheet_name='完成シフト')
                        worksheet = writer.sheets['完成シフト']
                        font_meiryo = Font(name='Meiryo')
                        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        align_center = Alignment(horizontal='center', vertical='center')
                        align_left   = Alignment(horizontal='left',  vertical='center')
                        fill_sat        = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
                        fill_sun        = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        fill_short      = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        fill_over       = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
                        fill_4days      = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                        fill_n3         = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
                        fill_n3_consec  = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                        fill_hope_off   = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        fill_hope_shift = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        for e in range(num_staff - 1, -1, -1):
                            worksheet.insert_rows(e + 3)
                        def staff_row(e): return (e + 1) * 2
                        sum_row_start = num_staff * 2 + 2
                        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                            for cell in row:
                                cell.font = font_meiryo; cell.border = border_thin
                                cell.alignment = align_left if cell.column == 1 else align_center
                        for c_idx, col_name in enumerate(cols):
                            if "土" in col_name:
                                for r_idx in range(1, worksheet.max_row + 1): worksheet.cell(row=r_idx, column=c_idx + 2).fill = fill_sat
                            elif "日" in col_name or "祝" in col_name:
                                for r_idx in range(1, worksheet.max_row + 1): worksheet.cell(row=r_idx, column=c_idx + 2).fill = fill_sun
                        for d, col_name in enumerate(cols):
                            actual_a = df_fin.loc[len(staff_names), col_name]
                            if actual_a != "":
                                if actual_a < day_req_list[d]: worksheet.cell(row=sum_row_start, column=d + 2).fill = fill_short
                                elif actual_a > day_req_list[d]: worksheet.cell(row=sum_row_start, column=d + 2).fill = fill_over
                        for e in range(num_staff):
                            xl_row = staff_row(e)
                            for d in range(num_days):
                                def is_d_work(day_idx, _e=e):
                                    if day_idx >= num_days: return False
                                    v = str(df_fin.loc[_e, cols[day_idx]])
                                    return v == 'A' or v == 'A残' or 'P' in v or 'Ｐ' in v
                                if (e, d) in hope_off_set:
                                    worksheet.cell(row=xl_row, column=d + 2).fill = fill_hope_off; continue
                                if (e, d) in hope_shift_dict:
                                    worksheet.cell(row=xl_row, column=d + 2).fill = fill_hope_shift; continue
                                if is_d_work(d) and is_d_work(d+1) and is_d_work(d+2) and is_d_work(d+3):
                                    for k in range(4): worksheet.cell(row=xl_row, column=d+k+2).fill = fill_4days
                                if d + 3 < num_days:
                                    if is_d_work(d) and is_d_work(d+1) and is_d_work(d+2) and str(df_fin.loc[e, cols[d+3]]) == 'D':
                                        for k in range(4): worksheet.cell(row=xl_row, column=d+k+2).fill = fill_n3
                                if d + 8 < num_days:
                                    if str(df_fin.loc[e, cols[d]]) == 'D' and str(df_fin.loc[e, cols[d+3]]) == 'D' and str(df_fin.loc[e, cols[d+6]]) == 'D':
                                        for k in range(9): worksheet.cell(row=xl_row, column=d+k+2).fill = fill_n3_consec

                        # ── ② 予定・実績シート ──
                        # スタッフごとに「予定」「実績」の2行を作成
                        plan_actual_rows = []
                        for e in range(num_staff):
                            # 予定行（アプリが自動入力）
                            row_plan = {"スタッフ名": staff_names[e], "区分": "予定"}
                            for d in range(num_days):
                                row_plan[cols[d]] = str(df_fin.loc[e, cols[d]])
                            plan_actual_rows.append(row_plan)
                            # 実績行（変更があった日だけ手入力。空欄＝予定通り）
                            row_actual = {"スタッフ名": staff_names[e], "区分": "実績"}
                            for d in range(num_days):
                                row_actual[cols[d]] = ""
                            plan_actual_rows.append(row_actual)

                        df_pa = pd.DataFrame(plan_actual_rows)
                        df_pa.to_excel(writer, index=False, sheet_name='予定・実績')
                        ws_pa = writer.sheets['予定・実績']

                        # 書式設定
                        fill_plan_hdr   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # 予定行 濃青
                        fill_actual_hdr = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # 実績行 緑
                        fill_plan_cell  = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")  # 予定セル 薄青
                        fill_actual_cell= PatternFill(start_color="F2F9EE", end_color="F2F9EE", fill_type="solid")  # 実績セル 薄緑（入力欄）
                        font_white      = Font(name='Meiryo', bold=True, color="FFFFFF")
                        font_normal     = Font(name='Meiryo')
                        font_gray       = Font(name='Meiryo', color="888888", italic=True)

                        # ヘッダー行
                        for cell in ws_pa[1]:
                            cell.fill = fill_plan_hdr; cell.font = font_white
                            cell.alignment = align_center; cell.border = border_thin
                        ws_pa.column_dimensions['A'].width = 14
                        ws_pa.column_dimensions['B'].width = 6
                        for col_idx in range(3, num_days + 3):
                            ws_pa.column_dimensions[ws_pa.cell(row=1, column=col_idx).column_letter].width = 5

                        # データ行の色分け
                        for row_idx in range(2, len(plan_actual_rows) + 2):
                            kubun_cell = ws_pa.cell(row=row_idx, column=2)
                            is_plan = (kubun_cell.value == "予定")
                            for col_idx in range(1, num_days + 3):
                                cell = ws_pa.cell(row=row_idx, column=col_idx)
                                cell.border = border_thin
                                cell.alignment = align_center if col_idx > 1 else align_left
                                if col_idx <= 2:
                                    # スタッフ名・区分列
                                    cell.fill = fill_plan_hdr if is_plan else fill_actual_hdr
                                    cell.font = font_white
                                elif is_plan:
                                    cell.fill = fill_plan_cell
                                    cell.font = font_normal
                                else:
                                    cell.fill = fill_actual_cell
                                    cell.font = font_gray
                                    if cell.value == "" or cell.value is None:
                                        cell.value = None  # 空欄を明示（入力欄として）

                        # 土日祝の列色（予定・実績どちらも）
                        for d, col_name in enumerate(cols):
                            col_idx = d + 3
                            col_letter = ws_pa.cell(row=1, column=col_idx).column_letter
                            if "土" in col_name:
                                ws_pa.cell(row=1, column=col_idx).fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                            elif "日" in col_name or "祝" in col_name:
                                ws_pa.cell(row=1, column=col_idx).fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

                        # 使い方メモを末尾に追記
                        note_row = len(plan_actual_rows) + 3
                        ws_pa.cell(row=note_row, column=1).value = "【記入方法】"
                        ws_pa.cell(row=note_row, column=1).font = Font(name='Meiryo', bold=True)
                        ws_pa.cell(row=note_row+1, column=1).value = "・実績行は変更があった日だけ入力してください（空欄＝予定通り）"
                        ws_pa.cell(row=note_row+2, column=1).value = "・入力例：急な欠勤→「公」、希望休変更→「A」など"
                        ws_pa.cell(row=note_row+3, column=1).value = "・月末にこのファイルをアップロードすると実績ベースで年間管理が更新されます"
                        for nr in range(note_row, note_row+4):
                            ws_pa.cell(row=nr, column=1).font = Font(name='Meiryo', color="555555")

                        # ── ③ 年間管理シート ──
                        df_report_for_annual, _, _, _ = build_review_report(df_fin, cols)
                        df_new_annual = build_annual_excel(df_report_for_annual)
                        df_new_annual.to_excel(writer, index=False, sheet_name="年間管理")
                        ws_annual = writer.sheets["年間管理"]
                        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        header_font = Font(name='Meiryo', bold=True, color="FFFFFF")
                        alert_fill  = PatternFill(start_color="FFD0D0", end_color="FFD0D0", fill_type="solid")
                        ok_fill     = PatternFill(start_color="D8F5D8", end_color="D8F5D8", fill_type="solid")
                        warn_fill   = PatternFill(start_color="FFF5CC", end_color="FFF5CC", fill_type="solid")
                        for cell in ws_annual[1]:
                            cell.fill = header_fill; cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        cols_annual = list(df_new_annual.columns)
                        def ac(col_name_a):
                            return cols_annual.index(col_name_a) + 1 if col_name_a in cols_annual else None
                        for row_idx in range(2, num_staff + 2):
                            for col_name_a, threshold, red_ok in [
                                ("年間義務残日数", 0, True),
                                ("有休残日数", 1, False),
                                ("前月公休実績", 8, False),
                            ]:
                                ci = ac(col_name_a)
                                if ci is None: continue
                                val = ws_annual.cell(row=row_idx, column=ci).value or 0
                                fval = float(val)
                                if red_ok and fval > threshold:
                                    ws_annual.cell(row=row_idx, column=ci).fill = alert_fill
                                elif not red_ok and fval <= threshold:
                                    ws_annual.cell(row=row_idx, column=ci).fill = warn_fill
                                elif red_ok:
                                    ws_annual.cell(row=row_idx, column=ci).fill = ok_fill
                        ws_annual.column_dimensions['A'].width = 14

                    st.download_button(
                        label=f"📥 シフト管理ファイルをダウンロード（パターン {i + 1}・{target_year}/{target_month:02d}）",
                        data=output.getvalue(),
                        file_name=f"シフト管理_{target_year}{target_month:02d}_パターン{i + 1}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_btn_{i}"
                    )
                    st.caption("📋 ファイルの中身：① 完成シフト（印刷用）　② 予定・実績（月中に実績を手入力）　③ 年間管理（来月引き継ぎ用）")

    except Exception as e:
        st.error(f"⚠️ エラーが発生しました: エクセルの形式が間違っているか、空白の行があります。({e})")
